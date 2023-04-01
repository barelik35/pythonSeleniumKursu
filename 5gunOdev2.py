from time import sleep
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import datetime
from datetime import date
import openpyxl 
from globalConstants import GlobalConstants as GC

#from selenium.common.exceptions import NoSuchElementException
class Test_Saucedemo:
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        now = datetime.now()
        self.folderPath = str(now.strftime("%d-%b-%Y")) #str(date.today())
        self.testTime=str(now.strftime("%H.%M")) #.%S
        Path(self.folderPath).mkdir(exist_ok=True)
        self.driver.get(GC.mainURL)

    def teardown_method(self):
        self.driver.quit()

    def waitForElementVisible(self,locator,timeout=10):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))
    def getData():
        excelFile = openpyxl.load_workbook("data/invalid_login.xlsx")
        selectedSheet=excelFile["SauceDemoLoginErrors"]
        totalRows=selectedSheet.max_row
        data=[]
        for i in range (2, totalRows+1):
            #if selectedSheet.cell(i,1).value is None:
            #    username=""
            #else:
            #     username= selectedSheet.cell(i,1).value
            username = selectedSheet.cell(i, 1).value or "" # Python da None değeri zaten bir boolean context içinde False olarak yorumlanır. 
            password= selectedSheet.cell(i,2).value or ""
            errorMessage= selectedSheet.cell(i,3).value or ""
            turpleData=(username,password,errorMessage)
            data.append(turpleData)
        data.append(turpleData)
        return data # [("","","Epic sadface: Username is required"),("1","","Epic sadface: Password is required"),("locked_out_user","secret_sauce","Epic sadface: Sorry, this user has been locked out.")]#data #, ("kullaniciadim","sifrem","errorMessage")

    def send(self,key,objectId):
        self.waitForElementVisible((By.ID,objectId))
        search_box = self.driver.find_element(By.ID,objectId)
        search_box.clear()
        search_box.send_keys(key)

    def loginclick(self):
        self.waitForElementVisible((By.ID,GC.loginButtonId))
        self.driver.find_element(By.ID,GC.loginButtonId).click()

    def errorMessageWeb(self):
        self.waitForElementVisible((By.XPATH,GC.errorMessageContainerXPATH))
        errorMessageContainer= self.driver.find_element(By.XPATH,GC.errorMessageContainerXPATH)
        return errorMessageContainer.text
    
    def standard_login(self):
        self.send("standard_user",GC.userNameId)
        self.send("secret_sauce",GC.passwordId)
        self.driver.find_element(By.ID,"login-button").click()

    def error_login(self):
        self.send("",GC.userNameId)
        self.send("",GC.passwordId)
        self.driver.find_element(By.ID,"login-button").click()

    def menuAction(self,action):
        self.waitForElementVisible((By.ID,GC.menuButton))
        menu= self.driver.find_element(By.ID,GC.menuButton)
        menu.click()
        self.waitForElementVisible((By.ID,action))
        actionLink= self.driver.find_element(By.ID,action)
        actionLink.click()
    def cartpage(self):
        self.waitForElementVisible((By.CLASS_NAME,"shopping_cart_link"))
        shoppingCart= self.driver.find_element(By.CLASS_NAME,"shopping_cart_link")
        shoppingCart.click()


    @pytest.mark.parametrize("username,password,errorMessageExcel",getData())
    def test_error_login(self,username,password,errorMessageExcel):
        self.send(username,str(GC.userNameId))
        self.send(password,str(GC.passwordId))
        self.loginclick()
        errorMessageWeb = self.errorMessageWeb()
        self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-sauce-login-error-{username}-{password}.png")
        assert errorMessageExcel == errorMessageWeb

    def test_standardUserInventory(self):
        self.standard_login()
        self.waitForElementVisible((By.XPATH,GC.inventoryContainerXPATH))
        self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-sauce-standardUserInventory.png")
        assert self.driver.current_url== GC.inventoryURL

    def test_numerOfItems(self):
        self.standard_login()
        self.waitForElementVisible((By.CLASS_NAME,GC.inventoryItem))
        products= list(self.driver.find_elements(By.CLASS_NAME,GC.inventoryItem))
        self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-sauce-numerOfItems.png")
        assert len(products)==6

    def test_nullIcon(self):
        self.error_login()
        self.waitForElementVisible((By.XPATH,GC.errorMessageContainerXPATH))
        icon1 = self.driver.find_elements(By.CSS_SELECTOR, ".form_group:nth-child(1) > .svg-inline--fa")
        icon2 = self.driver.find_elements(By.CSS_SELECTOR, ".form_group:nth-child(2) > .svg-inline--fa")
        self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-sauce-nullIcon.png")
        assert (len(icon1) & len(icon2)) > 0

    def test_logout(self):
        self.standard_login()
        self.menuAction("logout_sidebar_link")
        self.waitForElementVisible((By.ID,GC.userNameId))
        self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-sauce-logout.png")
        assert self.driver.current_url==GC.mainURL

    def test_cart(self):
        self.standard_login()
        self.cartpage()
        self.waitForElementVisible((By.ID,"cart_contents_container"))
        self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-sauce-cart.png")
        assert self.driver.current_url==GC.saucelabs
        
    @pytest.mark.xfail(strict=True, reason="blocked")
    def test_twitter(self):
        self.standard_login()
        self.waitForElementVisible((By.CLASS_NAME,"social_twitter"))
        twitter= self.driver.find_element(By.XPATH,"//*[@id=\"page_wrapper\"]/footer/ul/li[1]/a")
        twitter.click()
        self.driver.implicitly_wait(5)
        self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-sauce-twitter.png")
        assert self.driver.current_url=="https://twitter.com/saucelabs"
